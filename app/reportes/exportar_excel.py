import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def exportar_reporte_excel(filas, resumen):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    ws["A1"] = "NAHARA — Reporte de Ventas e Ingresos"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    encabezados = ["Fecha", "Pedido", "Producto", "Categoría", "Canal", "Cantidad", "Precio Unit. (Bs.)", "Subtotal (Bs.)"]
    fila_encabezado = 4
    for col, encabezado in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila_encabezado, column=col, value=encabezado)
        celda.font = Font(bold=True, color="F4CA00")
        celda.fill = PatternFill(start_color="111111", end_color="111111", fill_type="solid")

    fila = fila_encabezado + 1
    for f in filas:
        ws.cell(row=fila, column=1, value=f["fecha"].strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=fila, column=2, value=f["pedido_id"])
        ws.cell(row=fila, column=3, value=f["producto"])
        ws.cell(row=fila, column=4, value=f["categoria"])
        ws.cell(row=fila, column=5, value=f["canal"])
        ws.cell(row=fila, column=6, value=f["cantidad"])
        ws.cell(row=fila, column=7, value=float(f["precio_unitario"]))
        ws.cell(row=fila, column=8, value=float(f["subtotal"]))
        fila += 1

    fila += 1
    ws.cell(row=fila, column=1, value="Total unidades:").font = Font(bold=True)
    ws.cell(row=fila, column=2, value=resumen["total_unidades"])
    fila += 1
    ws.cell(row=fila, column=1, value="Total ingresos (Bs.):").font = Font(bold=True)
    ws.cell(row=fila, column=2, value=float(resumen["total_ingresos"]))
    fila += 2

    ws.cell(row=fila, column=1, value="Canal").font = Font(bold=True)
    ws.cell(row=fila, column=2, value="Unidades").font = Font(bold=True)
    ws.cell(row=fila, column=3, value="Ingresos (Bs.)").font = Font(bold=True)
    for canal, datos_canal in resumen["por_canal"].items():
        fila += 1
        ws.cell(row=fila, column=1, value=canal)
        ws.cell(row=fila, column=2, value=datos_canal["unidades"])
        ws.cell(row=fila, column=3, value=float(datos_canal["ingresos"]))

    for columna, ancho in zip("ABCDEFGH", [18, 8, 28, 18, 12, 10, 16, 14]):
        ws.column_dimensions[columna].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
